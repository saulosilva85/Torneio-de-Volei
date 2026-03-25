import streamlit as st
import random
from openpyxl import Workbook
import io
import math

st.set_page_config(page_title="Torneio de Vôlei", layout="centered")

st.title("🏐 OPEN VILLAGE FLEX 🏐")

# Inputs (COM KEY ÚNICA)
st.markdown("## 🔹 Cabeças de Chave (1 por time)")
cabecas_input = st.text_area("Digite um nome por linha", key="cabecas")

st.markdown("## 🔹 Mulheres (mínimo = nº de times)")
mulheres_input = st.text_area("Digite um nome por linha", key="mulheres")

st.markdown("## 🔹 Demais Jogadores")
jogadores_input = st.text_area("Digite os demais jogadores", key="jogadores")


# -------------------------
# GERAR JOGOS DO GRUPO
# -------------------------
def gerar_jogos_grupo(times):
    jogos = []
    for i in range(len(times)):
        for j in range(i + 1, len(times)):
            jogos.append((times[i], times[j]))
    return jogos


# -------------------------
# DIVIDIR GRUPOS
# -------------------------
def dividir_grupos(times):
    random.shuffle(times)
    metade = math.ceil(len(times) / 2)
    return times[:metade], times[metade:]


# -------------------------
# GERAR EXCEL
# -------------------------
def gerar_excel(times, nomes_times, grupoA, grupoB, jogosA, jogosB):
    wb = Workbook()

    # ABA TIMES
    ws1 = wb.active
    ws1.title = "Times"

    row = 1
    for i, time in enumerate(times):
        ws1.cell(row=row, column=1, value=nomes_times[i])
        row += 1
        for jogador in time:
            ws1.cell(row=row, column=2, value=jogador)
            row += 1
        row += 1

    # ABA GRUPOS
    ws2 = wb.create_sheet("Grupos")

    ws2["A1"] = "Grupo A"
    for i, t in enumerate(grupoA, start=2):
        ws2[f"A{i}"] = t

    ws2["C1"] = "Grupo B"
    for i, t in enumerate(grupoB, start=2):
        ws2[f"C{i}"] = t

    # ABA JOGOS
    ws3 = wb.create_sheet("Jogos")

    ws3["A1"] = "Grupo A"
    for i, j in enumerate(jogosA, start=2):
        ws3.cell(row=i, column=1, value=j[0])
        ws3.cell(row=i, column=2, value="vs")
        ws3.cell(row=i, column=3, value=j[1])

    ws3["E1"] = "Grupo B"
    for i, j in enumerate(jogosB, start=2):
        ws3.cell(row=i, column=5, value=j[0])
        ws3.cell(row=i, column=6, value="vs")
        ws3.cell(row=i, column=7, value=j[1])

    # ABA MATA-MATA
    ws4 = wb.create_sheet("Mata-Mata")

    ws4["A1"] = "Semi 1"
    ws4["A2"] = "1º A"
    ws4["A3"] = "2º B"

    ws4["C1"] = "Semi 2"
    ws4["C2"] = "1º B"
    ws4["C3"] = "2º A"

    ws4["E1"] = "Final"
    ws4["E2"] = "Vencedor Semi 1"
    ws4["E3"] = "Vencedor Semi 2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


# -------------------------
# BOTÃO
# -------------------------
if st.button("🎲 Sortear Torneio"):

    cabecas = [n.strip() for n in cabecas_input.split("\n") if n.strip()]
    mulheres = [n.strip() for n in mulheres_input.split("\n") if n.strip()]
    jogadores = [n.strip() for n in jogadores_input.split("\n") if n.strip()]

    if len(cabecas) < 2:
        st.error("Informe pelo menos 2 cabeças de chave.")
        st.stop()

    if len(mulheres) < len(cabecas):
        st.error("Número de mulheres deve ser igual ou maior que o número de times.")
        st.stop()

    todos = cabecas + mulheres + jogadores
    if len(todos) != len(set(todos)):
        st.error("Existem nomes duplicados.")
        st.stop()

    random.shuffle(mulheres)
    random.shuffle(jogadores)

    # Criar times
    times = []
    nomes_times = []

    for i in range(len(cabecas)):
        time = [cabecas[i], mulheres[i]]
        times.append(time)
        nomes_times.append(f"Time {i+1}")

    # Distribuir jogadores
    i = 0
    while jogadores:
        times[i % len(times)].append(jogadores.pop(0))
        i += 1

    # Exibir times
    st.markdown("## 🏆 Times")
    for i, time in enumerate(times):
        st.markdown(f"### {nomes_times[i]}")
        for jogador in time:
            st.write(f"• {jogador}")

    # Grupos
    grupoA, grupoB = dividir_grupos(nomes_times.copy())

    st.markdown("## 🔵 Grupo A")
    for t in grupoA:
        st.write(t)

    st.markdown("## 🔴 Grupo B")
    for t in grupoB:
        st.write(t)

    # Jogos
    jogosA = gerar_jogos_grupo(grupoA)
    jogosB = gerar_jogos_grupo(grupoB)

    st.markdown("## 📅 Jogos - Grupo A")
    for j in jogosA:
        st.write(f"{j[0]} 🆚 {j[1]}")

    st.markdown("## 📅 Jogos - Grupo B")
    for j in jogosB:
        st.write(f"{j[0]} 🆚 {j[1]}")

    # Mata-mata
    st.markdown("## 🏆 Mata-Mata")
    st.write("1º A 🆚 2º B")
    st.write("1º B 🆚 2º A")
    st.write("Final")

    # Excel
    excel = gerar_excel(times, nomes_times, grupoA, grupoB, jogosA, jogosB)

    st.download_button(
        label="📊 Baixar Excel",
        data=excel,
        file_name="torneio_volei_flex.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )